import os
import io
import uuid
import tempfile
import logging
from pathlib import Path
from typing import Optional

import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="PDF to Excel Converter", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = Path(tempfile.gettempdir()) / "pdf_uploads"
OUTPUT_DIR = Path(tempfile.gettempdir()) / "excel_outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# ─── Extraction helpers ───────────────────────────────────────────────────────

def has_selectable_text(pdf_path: str, max_pages: int = 2) -> bool:
    """Fast check for digital PDFs so scanned PDFs can skip slow table parsing."""
    try:
        import pypdfium2 as pdfium
    except ImportError:
        return True

    try:
        pdf = pdfium.PdfDocument(pdf_path)
        for page_num in range(min(len(pdf), max_pages)):
            text = pdf[page_num].get_textpage().get_text_bounded().strip()
            if text:
                return True
        return False
    except Exception:
        logger.exception("Fast selectable-text check failed - falling back to pdfplumber.")
        return True


def extract_tables_pdfplumber(pdf_path: str) -> list[pd.DataFrame]:
    """Extract tables from a native/digital PDF using pdfplumber.
    All pages are merged into a SINGLE DataFrame to avoid multi-sheet output."""
    all_rows = []
    detected_header = None

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            logger.info(f"  pdfplumber scanning page {page_num}/{len(pdf.pages)}")

            table_settings_strict = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
                "edge_min_length": 3,
                "min_words_vertical": 1,
                "min_words_horizontal": 1,
            }
            extracted = page.extract_tables(table_settings_strict)

            if not extracted:
                table_settings_stream = {
                    "vertical_strategy": "text",
                    "horizontal_strategy": "text",
                    "snap_tolerance": 5,
                    "join_tolerance": 5,
                }
                extracted = page.extract_tables(table_settings_stream)

            if not extracted:
                extracted = page.extract_tables()

            for tbl in extracted:
                if not tbl or len(tbl) < 2:
                    continue

                # Detect header: use first page's first row as the global header
                first_row = tbl[0]
                if detected_header is None:
                    detected_header = first_row
                    data_rows = tbl[1:]
                else:
                    # Skip repeated header rows on subsequent pages
                    if tbl[0] == detected_header or all(
                        str(c).strip().lower() == str(h).strip().lower()
                        for c, h in zip(tbl[0], detected_header)
                    ):
                        data_rows = tbl[1:]
                    else:
                        data_rows = tbl  # no header on this page chunk

                n_cols = len(detected_header)
                for row in data_rows:
                    if len(row) < n_cols:
                        row = row + [""] * (n_cols - len(row))
                    all_rows.append(row[:n_cols])

    if not all_rows or detected_header is None:
        return []

    df = pd.DataFrame(all_rows, columns=detected_header)
    df = df.replace("", pd.NA).dropna(how="all").fillna("")
    if df.empty:
        return []
    return [df]  # single DataFrame — one sheet


def extract_text_rows_pdfplumber(pdf_path: str) -> list[pd.DataFrame]:
    """Fallback for text PDFs with no detectable table grid.
    All pages merged into a SINGLE DataFrame."""
    all_rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=2, y_tolerance=4) or ""
            rows = [[line.strip()] for line in text.splitlines() if line.strip()]
            if rows:
                logger.info(f"  pdfplumber text fallback found {len(rows)} row(s) on page {page_num}")
                all_rows.extend(rows)
    if not all_rows:
        return []
    return [pd.DataFrame(all_rows, columns=[""])]


def ocr_image_to_rows(image) -> list[list[str]]:
    """Use Tesseract word positions to rebuild simple row/column layout."""
    import pytesseract
    from pytesseract import Output
    import numpy as np

    data = pytesseract.image_to_data(
        image,
        config="--psm 11",
        output_type=Output.DICT,
    )

    words = []
    for idx, text in enumerate(data.get("text", [])):
        text = str(text).strip().strip("|{}")
        try:
            conf = float(data["conf"][idx])
        except (TypeError, ValueError):
            conf = -1

        if not text or conf < 25:
            continue

        left = int(data["left"][idx])
        top = int(data["top"][idx])
        width = int(data["width"][idx])
        height = int(data["height"][idx])
        words.append(
            {
                "text": text,
                "left": left,
                "right": left + width,
                "top": top,
                "center_x": left + width / 2,
                "center_y": top + height / 2,
            }
        )

    if not words:
        return []

    rows = []
    for word in sorted(words, key=lambda item: item["center_y"]):
        for row in rows:
            if abs(row["center_y"] - word["center_y"]) <= 14:
                row["words"].append(word)
                row["center_y"] = sum(item["center_y"] for item in row["words"]) / len(row["words"])
                break
        else:
            rows.append({"center_y": word["center_y"], "words": [word]})

    separators = detect_vertical_separators(image, np)
    structured_rows = []
    for row in sorted(rows, key=lambda item: item["center_y"]):
        row_words = sorted(row["words"], key=lambda item: item["left"])

        if separators:
            cells = [""] * (len(separators) + 1)
            for word in row_words:
                cell_index = 0
                for separator in separators:
                    if word["center_x"] > separator:
                        cell_index += 1
                    else:
                        break

                cells[cell_index] = (cells[cell_index] + " " + word["text"]).strip()

            while cells and not cells[-1]:
                cells.pop()

            if any(cells):
                structured_rows.append(cells)
            continue

        cells = []
        current_words = []
        previous_right = None

        for word in row_words:
            if previous_right is not None and word["left"] - previous_right > 45:
                cells.append(" ".join(current_words).strip())
                current_words = []

            current_words.append(word["text"])
            previous_right = word["right"]

        if current_words:
            cells.append(" ".join(current_words).strip())

        if any(cells):
            structured_rows.append(cells)

    return structured_rows


def detect_vertical_separators(image, np_module) -> list[int]:
    """Detect strong vertical table lines and return internal separator x positions."""
    arr = np_module.array(image)
    dark_pixels = arr < 80
    col_counts = dark_pixels.sum(axis=0)
    height, width = arr.shape
    candidate_cols = [idx for idx, count in enumerate(col_counts) if count > height * 0.25]

    groups = []
    for col in candidate_cols:
        if not groups or col - groups[-1][-1] > 3:
            groups.append([])
        groups[-1].append(col)

    centers = [round(sum(group) / len(group)) for group in groups if group]
    internal = [center for center in centers if width * 0.10 < center < width * 0.90]
    return internal


def extract_tables_ocr(pdf_path: str) -> list[pd.DataFrame]:
    """
    Fallback OCR path for scanned PDFs.
    Uses pdf2image + easyocr to extract text, then heuristically rebuilds tables.
    """
    try:
        import pypdfium2 as pdfium
        from PIL import ImageOps
    except ImportError:
        logger.warning("pypdfium2 or Pillow not installed - skipping Tesseract OCR path.")
    else:
        try:
            pdf = pdfium.PdfDocument(pdf_path)
            tables = []
            for page_num in range(len(pdf)):
                logger.info(f"  Tesseract OCR scanning page {page_num + 1}/{len(pdf)}")
                image = pdf[page_num].render(scale=3).to_pil().convert("L")
                image = ImageOps.autocontrast(image)
                rows = ocr_image_to_rows(image)
                if rows:
                    n_cols = max(len(row) for row in rows)
                    padded = [row + [""] * (n_cols - len(row)) for row in rows]
                    tables.append(pd.DataFrame(padded, columns=[""] * n_cols))

            if tables:
                return tables
        except Exception:
            logger.exception("Tesseract OCR extraction failed - trying optional OCR path.")

    try:
        from pdf2image import convert_from_path
        import easyocr
    except ImportError:
        logger.warning("pdf2image or easyocr not installed – skipping OCR path.")
        return []

    try:
        reader = easyocr.Reader(["en"], gpu=False)
        tables = []

        images = convert_from_path(pdf_path, dpi=200)
        for page_num, img in enumerate(images, start=1):
            logger.info(f"  OCR scanning page {page_num}/{len(images)}")
            import numpy as np
            img_array = np.array(img)
            results = reader.readtext(img_array, detail=1, paragraph=False)

            if not results:
                continue

            # Sort by vertical position, then horizontal
            results_sorted = sorted(results, key=lambda r: (round(r[0][0][1] / 15), r[0][0][0]))

            # Group into rows by y-coordinate proximity
            rows: dict[int, list] = {}
            for bbox, text, conf in results_sorted:
                y_center = int((bbox[0][1] + bbox[2][1]) / 2)
                row_key = y_center // 20  # 20-pixel row bucket
                rows.setdefault(row_key, []).append((bbox[0][0], text))

            structured = []
            for key in sorted(rows.keys()):
                row_items = sorted(rows[key], key=lambda x: x[0])
                structured.append([item[1] for item in row_items])

            if len(structured) >= 2:
                n_cols = max(len(r) for r in structured)
                padded = [r + [""] * (n_cols - len(r)) for r in structured]
                df = pd.DataFrame(padded[1:], columns=padded[0])
                df = df.replace("", pd.NA).dropna(how="all").fillna("")
                if not df.empty:
                    tables.append(df)

        return tables
    except Exception:
        logger.exception("OCR extraction failed - continuing without OCR results.")
        return []


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean and normalize a DataFrame for Excel output."""
    # Strip whitespace from string cells
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
    # Drop fully empty rows and columns
    df = df.dropna(how="all").reset_index(drop=True)
    df = df.loc[:, df.notna().any()]
    # Replace None with empty string
    df = df.fillna("")
    return df


def write_excel(tables: list[pd.DataFrame], output_path: str) -> None:
    """Write ALL extracted DataFrames into a SINGLE sheet, stacked vertically."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    if not tables:
        ws["A1"] = "No tables were detected in the uploaded PDF."
        wb.save(output_path)
        return

    current_row = 1

    for table_index, table in enumerate(tables):
        table = table.copy().fillna("")

        if table_index > 0:
            current_row += 1  # blank separator row between tables

        col_names = [str(c).strip() for c in table.columns]
        has_real_header = any(
            c and c not in ("", "0", "1", "2", "3", "4", "5", "None")
            for c in col_names
        )

        if has_real_header:
            for col_num, col_name in enumerate(col_names, start=1):
                ws.cell(row=current_row, column=col_num, value=col_name)
            current_row += 1

        for row_tuple in table.itertuples(index=False):
            for col_num, value in enumerate(row_tuple, start=1):
                cell_val = str(value) if value not in ("", None) else ""
                ws.cell(row=current_row, column=col_num, value=cell_val)
            current_row += 1  # increment after EVERY row

    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 60)

    wb.save(output_path)


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    html_path = Path(__file__).parent / "index.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>PDF to Excel Converter API</h1><p>Upload a PDF to /convert</p>")


@app.post("/convert")
async def convert_pdf(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")

    job_id    = uuid.uuid4().hex
    pdf_path  = str(UPLOAD_DIR / f"{job_id}.pdf")
    xlsx_name = Path(file.filename).stem + "_converted.xlsx"
    xlsx_path = str(OUTPUT_DIR / f"{job_id}.xlsx")

    # Save uploaded PDF
    content = await file.read()
    with open(pdf_path, "wb") as f:
        f.write(content)
    logger.info(f"Saved PDF → {pdf_path}")

    try:
        digital_pdf = has_selectable_text(pdf_path)

        # Phase 1: attempt native extraction
        logger.info("Phase 1: pdfplumber extraction …")
        if digital_pdf:
            tables = extract_tables_pdfplumber(pdf_path)
        else:
            logger.info("Phase 1: scanned PDF detected - skipping pdfplumber table scan.")
            tables = []

        # Phase 2: fallback to OCR if nothing found
        if not tables:
            logger.info("Phase 2: No native tables found – trying OCR …")
            tables = extract_text_rows_pdfplumber(pdf_path) if digital_pdf else []
            if not tables:
                logger.info("Phase 3: trying OCR ...")
                tables = extract_tables_ocr(pdf_path)

        if not tables:
            raise HTTPException(
                status_code=422,
                detail="No readable text could be detected in this PDF. "
                       "Please try a clearer scan or a higher-resolution PDF.",
            )

        # Phase 3: clean & write
        logger.info(f"Cleaning {len(tables)} table(s) …")
        tables = [clean_dataframe(df) for df in tables if not df.empty]
        write_excel(tables, xlsx_path)
        logger.info(f"Excel saved → {xlsx_path}")

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("PDF conversion failed.")
        raise HTTPException(
            status_code=500,
            detail=f"Conversion failed: {exc}",
        ) from exc
    finally:
        # Always clean up the uploaded PDF
        try:
            os.remove(pdf_path)
        except OSError:
            pass

    return FileResponse(
        path=xlsx_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=xlsx_name,
        background=None,
    )


@app.get("/health")
async def health():
    return {"status": "ok", "service": "PDF to Excel Converter"}
